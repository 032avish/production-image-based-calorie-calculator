import openpyxl  #to read write excel
import requests   #to make http request to google
from bs4 import BeautifulSoup    #to pull data from html
import os   #various os task

api_key = "AIzaSyAS3t0GV75JbPg6yIfeQ2vcHxexA6cKVXM"
engine_id = "2668a33ec270e4242"


def search_image_url(name, api_key, engine_id):
    try:
        # Construct the search query
        query = f"{name} image"
        search_url = f"https://www.googleapis.com/customsearch/v1?q={query}&cx={engine_id}&searchType=image&key={api_key}"
        
        # Send GET request to the API
        response = requests.get(search_url)
        data = response.json()
        
        # Extract image URL from the response
        if 'items' in data and len(data['items']) > 0:
            image_url = data['items'][0]['link']
            return image_url
        else:
            print(f"No image found for {name}")
            return None
        
    except Exception as e:
        print(f"An error occurred while searching for the image of '{name}': {str(e)}")
        return None


def download_and_save_image(url, name):
    # try:
    #     response = requests.get(url)
    #     if response.status_code == 200:
    #         # Extracting file extension from URL
    #         extension = url.split('.')[-1]
    #         # Sanitize the name to remove invalid characters
    #         sanitized_name = ''.join(char for char in name if char.isalnum())
    #         # Setting the filename
    #         filename = f"{sanitized_name}.{extension}"
    #         # Saving the image to current directory
    #         with open(filename, 'wb') as f:
    #             f.write(response.content)
    #         print(f"Image '{filename}' downloaded successfully.")
    #         return filename
    #     else:
    #         print(f"Failed to download image: {response.status_code}")
    #         return None

    # except Exception as e:
    #     print(f"An error occurred while downloading image from '{url}': {str(e)}")
    #     return None
    return url

def download_and_display_images(file_path):
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        # Select the active worksheet
        ws = wb.active
        
        # Iterate over rows and columns to read data
        for row in ws.iter_rows(min_row=1, max_col=1, max_row=ws.max_row):
            name = row[0].value
            print("Name:", name)

            # Search for image URL
            image_url = search_image_url(name, api_key, engine_id)
            if image_url:
                print("Image URL:", image_url)
                # Download and save the image
                image_path = download_and_save_image(image_url, name)
                if image_path:
                   # Write image path to the next cell
                   next_cell = ws.cell(row=row[0].row, column=row[0].column + 1)
                   try:
                     next_cell.value = image_url
                     print(f"Image URL '{image_url}' written to Excel successfully.")
                   except Exception as e:
                     print(f"An error occurred while writing image URL to Excel: {str(e)}")


        # Save the workbook
        wb.save(file_path)

    except FileNotFoundError:
        print(f"File '{file_path}' not found.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    file_path = "C:\\Users\\avish\\OneDrive\\Desktop\\test\\Book2.xlsx"
    download_and_display_images(file_path)